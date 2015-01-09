from openpyxl import load_workbook



wb2 = load_workbook('D://Expenses01.xlsx', data_only = True)

print wb2.get_sheet_names()
ws4 = wb2.get_sheet_by_name("Sheet1")

rowCount = len(ws4.rows) ## Get the row Count
colCount = len(ws4.columns) ## Get the column count
for i in range(1, rowCount+1):
    for j in range(1, colCount+1):
        sVal = ws4.cell(row = i, column = j)
        print sVal.value
##        print sVal.data_type
##        if sVal.data_type == "f":
##            print "in if contidition"
##            print sVal.internal_value
##        else:
##            print sVal.value


